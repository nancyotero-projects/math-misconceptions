"""Prepare remediation experiment datasets from the recommendation workbook.

Run from the repository root:

    python remediation-experiments/scripts/prepare_dataset.py

Outputs:
    remediation-experiments/data/processed/recommendations_long.csv
    remediation-experiments/data/processed/remediation_cases.jsonl
    remediation-experiments/data/splits/pilot_20.json
    remediation-experiments/data/splits/full_55.json
"""

from __future__ import annotations

import csv
import json
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


EXPERIMENT_ROOT = Path(__file__).resolve().parents[1]
REPO_ROOT = Path(__file__).resolve().parents[2]
SOURCE_XLSX = EXPERIMENT_ROOT / "data" / "raw" / "recommendation_dataset_original.xlsx"
MAE_DATA_JSON = REPO_ROOT / "data" / "data.json"
PROCESSED_DIR = EXPERIMENT_ROOT / "data" / "processed"
SPLITS_DIR = EXPERIMENT_ROOT / "data" / "splits"


def clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def load_topic_lookup() -> dict[str, str]:
    if not MAE_DATA_JSON.exists():
        return {}

    with MAE_DATA_JSON.open("r", encoding="utf-8") as f:
        rows = json.load(f)

    lookup: dict[str, str] = {}
    for row in rows:
        misconception_id = clean(row.get("Misconception ID"))
        topic = clean(row.get("Topic"))
        if misconception_id and topic:
            lookup[misconception_id] = topic
    return lookup


def workbook_rows() -> list[dict[str, str]]:
    workbook = load_workbook(SOURCE_XLSX, data_only=True)
    sheet = workbook.active
    headers = [clean(cell.value) for cell in sheet[1]]
    rows = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        record = {headers[i]: clean(value) for i, value in enumerate(row)}
        if record.get("ID"):
            rows.append(record)

    return rows


def build_long_rows(source_rows: list[dict[str, str]], topic_lookup: dict[str, str]) -> list[dict[str, str]]:
    long_rows = []

    for source in source_rows:
        misconception_id = source["ID"]
        topic = topic_lookup.get(misconception_id, "")

        training_examples = []
        for example_number in range(1, 5):
            training_examples.append(
                {
                    "example_number": example_number,
                    "training_example": source.get(f"Correct training example {example_number}", ""),
                    "training_explanation": source.get(
                        f"Explanation of presented MaE in example {example_number}", ""
                    ),
                }
            )

        for recommendation_number in range(1, 5):
            case_id = f"{misconception_id}_rec{recommendation_number}"
            long_rows.append(
                {
                    "case_id": case_id,
                    "misconception_id": misconception_id,
                    "topic": topic,
                    "misconception": source["Misconception"],
                    "recommendation_number": str(recommendation_number),
                    "research_recommendation": source.get(f"Suggestion {recommendation_number}", ""),
                    "gold_activity": source.get(f"Suggested Example {recommendation_number}", ""),
                    "gold_correct_response": source.get(f"Correct A {recommendation_number}", ""),
                    "gold_incorrect_response": source.get(f"Incorrect A {recommendation_number}", ""),
                    "training_examples_json": json.dumps(training_examples, ensure_ascii=False),
                }
            )

    return long_rows


def build_cases(long_rows: list[dict[str, str]]) -> list[dict[str, object]]:
    grouped: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in long_rows:
        grouped[row["misconception_id"]].append(row)

    cases = []
    for misconception_id in sorted(grouped):
        rows = sorted(grouped[misconception_id], key=lambda item: int(item["recommendation_number"]))
        first = rows[0]
        training_examples = json.loads(first["training_examples_json"])

        student_work_pattern = []
        for row in rows:
            if row["gold_activity"] and row["gold_incorrect_response"]:
                student_work_pattern.append(
                    {
                        "activity_or_question": row["gold_activity"],
                        "student_incorrect_response": row["gold_incorrect_response"],
                    }
                )

        cases.append(
            {
                "case_id": misconception_id,
                "misconception_id": misconception_id,
                "topic": first["topic"],
                "misconception": first["misconception"],
                "student_work_pattern": student_work_pattern,
                "training_examples": training_examples,
                "research_recommendations": [
                    {
                        "recommendation_number": int(row["recommendation_number"]),
                        "research_recommendation": row["research_recommendation"],
                        "gold_activity": row["gold_activity"],
                        "gold_correct_response": row["gold_correct_response"],
                        "gold_incorrect_response": row["gold_incorrect_response"],
                    }
                    for row in rows
                ],
            }
        )

    return cases


def write_csv(rows: list[dict[str, str]], path: Path) -> None:
    fieldnames = list(rows[0].keys())
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_jsonl(rows: list[dict[str, object]], path: Path) -> None:
    with path.open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")


def write_splits(cases: list[dict[str, object]]) -> None:
    full_ids = [str(case["case_id"]) for case in cases]

    # Stratified-ish pilot: favor known difficult areas while keeping topic spread.
    preferred_ids = [
        "MaE01",
        "MaE02",
        "MaE03",
        "MaE06",
        "MaE10",
        "MaE14",
        "MaE18",
        "MaE23",
        "MaE24",
        "MaE25",
        "MaE26",
        "MaE28",
        "MaE31",
        "MaE35",
        "MaE37",
        "MaE40",
        "MaE43",
        "MaE45",
        "MaE49",
        "MaE52",
    ]
    pilot_ids = [case_id for case_id in preferred_ids if case_id in full_ids]

    (SPLITS_DIR / "pilot_20.json").write_text(
        json.dumps({"split_name": "pilot_20", "case_ids": pilot_ids}, indent=2) + "\n",
        encoding="utf-8",
    )
    (SPLITS_DIR / "full_55.json").write_text(
        json.dumps({"split_name": "full_55", "case_ids": full_ids}, indent=2) + "\n",
        encoding="utf-8",
    )


def main() -> None:
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    SPLITS_DIR.mkdir(parents=True, exist_ok=True)

    topic_lookup = load_topic_lookup()
    source_rows = workbook_rows()
    long_rows = build_long_rows(source_rows, topic_lookup)
    cases = build_cases(long_rows)

    write_csv(long_rows, PROCESSED_DIR / "recommendations_long.csv")
    write_jsonl(cases, PROCESSED_DIR / "remediation_cases.jsonl")
    write_splits(cases)

    print(f"Wrote {len(long_rows)} long recommendation rows.")
    print(f"Wrote {len(cases)} remediation cases.")


if __name__ == "__main__":
    main()
